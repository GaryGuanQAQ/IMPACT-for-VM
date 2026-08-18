"""Copy thorax DICOM studies into condition-specific folders.

Run this script through RUN_THORAX_DICOM_ORGANIZER.bat on the Windows VM.
The spreadsheet must contain columns named scan_dir and thorax_condition.
"""

from __future__ import annotations

import argparse
import csv
import re
import shutil
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path, PurePosixPath
from typing import Iterable


DEFAULT_SOURCE_ROOT = Path(r"Z:\sectra2")
DEFAULT_OUTPUT_ROOT = Path(r"Z:\thorax_data")
CONDITION_FOLDERS = {
    "pt": "pneumothorax",
    "pneumothorax": "pneumothorax",
    "ht": "haemothorax",
    "hemothorax": "haemothorax",
    "haemothorax": "haemothorax",
    "htpt_joint": "joint",
    "hpt_joint": "joint",
    "joint": "joint",
}
INVALID_SCAN_DIRS = {"", "#n/a", "n/a", "na", "error", "none", "null", "nan"}


def choose_spreadsheet() -> Path | None:
    """Show a file chooser when the launcher is double-clicked."""
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        selected = filedialog.askopenfilename(
            title="Select the thorax spreadsheet",
            filetypes=[
                ("Spreadsheet files", "*.xlsx *.xls *.csv"),
                ("Excel files", "*.xlsx *.xls"),
                ("CSV files", "*.csv"),
            ],
        )
        root.destroy()
        return Path(selected) if selected else None
    except Exception:
        return None


def normalise_column_name(name: object) -> str:
    return str(name).strip().lower().replace(" ", "_")


def read_rows(spreadsheet: Path) -> Iterable[dict[str, object]]:
    suffix = spreadsheet.suffix.lower()
    if suffix == ".csv":
        with spreadsheet.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                yield {normalise_column_name(key): value for key, value in row.items()}
        return

    if suffix not in {".xlsx", ".xls"}:
        raise ValueError("Use an .xlsx, .xls, or .csv spreadsheet.")

    try:
        import openpyxl
    except ImportError as error:
        raise RuntimeError(
            "Excel files require openpyxl. Run: python -m pip install openpyxl"
        ) from error

    if suffix == ".xls":
        raise ValueError(
            "Legacy .xls files are not supported directly. Save the spreadsheet as .xlsx, then run again."
        )

    workbook = openpyxl.load_workbook(spreadsheet, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise ValueError("The spreadsheet is empty.")
        headers = [normalise_column_name(cell) for cell in header_row]
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            yield dict(zip(headers, values))
    finally:
        workbook.close()


def normalise_scan_dir(value: object) -> PurePosixPath | None:
    text = str(value or "").strip().replace("\\", "/")
    if text.lower() in INVALID_SCAN_DIRS:
        return None

    parts = [part for part in text.split("/") if part and part not in {".", ".."}]
    if len(parts) < 2 or any(not re.fullmatch(r"\d+", part) for part in parts):
        return None
    return PurePosixPath(*parts)


def condition_folder(value: object) -> str | None:
    condition = str(value or "").strip().lower().replace("-", "_").replace(" ", "_")
    return CONDITION_FOLDERS.get(condition)


def copy_study(source_root: Path, output_root: Path, condition: str, scan_dir: PurePosixPath) -> None:
    relative_path = Path(*scan_dir.parts)
    source = source_root / relative_path
    destination = output_root / condition / relative_path
    if not source.is_dir():
        raise FileNotFoundError(source)
    shutil.copytree(source, destination, dirs_exist_ok=True)


def write_report(output_root: Path, counts: Counter, skipped: list[str]) -> Path:
    report_path = output_root / f"organize_report_{datetime.now():%Y%m%d_%H%M%S}.txt"
    lines = [
        "Thorax DICOM organization report",
        f"Created: {datetime.now():%Y-%m-%d %H:%M:%S}",
        "",
        "Summary:",
    ]
    lines.extend(f"  {name}: {count}" for name, count in sorted(counts.items()))
    if skipped:
        lines.extend(["", "Skipped rows:"])
        lines.extend(f"  {item}" for item in skipped)
    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return report_path


def organise(spreadsheet: Path, source_root: Path, output_root: Path) -> int:
    if not source_root.is_dir():
        print(f"ERROR: DICOM source folder was not found: {source_root}")
        print("Use --source-root to select the correct folder, for example Z:\\sectra.")
        return 2

    output_root.mkdir(parents=True, exist_ok=True)
    for folder in sorted(set(CONDITION_FOLDERS.values())):
        (output_root / folder).mkdir(exist_ok=True)

    counts: Counter = Counter()
    skipped: list[str] = []
    seen: set[tuple[str, PurePosixPath]] = set()

    for row_number, row in enumerate(read_rows(spreadsheet), start=2):
        raw_scan_dir = row.get("scan_dir")
        raw_condition = row.get("thorax_condition")
        scan_dir = normalise_scan_dir(raw_scan_dir)
        condition = condition_folder(raw_condition)

        if scan_dir is None:
            counts["skipped_invalid_scan_dir"] += 1
            skipped.append(f"Row {row_number}: invalid scan_dir ({raw_scan_dir!r})")
            continue
        if condition is None:
            counts["skipped_invalid_condition"] += 1
            skipped.append(f"Row {row_number}: invalid thorax_condition ({raw_condition!r})")
            continue
        key = (condition, scan_dir)
        if key in seen:
            counts["skipped_duplicate"] += 1
            continue
        seen.add(key)

        try:
            copy_study(source_root, output_root, condition, scan_dir)
            counts[f"copied_{condition}"] += 1
            print(f"Copied {scan_dir} -> {condition}")
        except FileNotFoundError:
            counts["skipped_missing_source"] += 1
            skipped.append(f"Row {row_number}: source not found ({scan_dir})")
            print(f"SKIPPED (source not found): {scan_dir}")

    report_path = write_report(output_root, counts, skipped)
    print("\nFinished.")
    print(f"Output folder: {output_root}")
    print(f"Report: {report_path}")
    print("Summary: " + ", ".join(f"{key}={value}" for key, value in sorted(counts.items())))
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Copy thorax DICOM studies by spreadsheet condition.")
    parser.add_argument("spreadsheet", nargs="?", type=Path, help="Path to the spreadsheet with scan_dir and thorax_condition.")
    parser.add_argument("--source-root", type=Path, default=DEFAULT_SOURCE_ROOT, help=r"DICOM root folder (default: Z:\sectra2).")
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT, help=r"Destination folder (default: Z:\thorax_data).")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    spreadsheet = args.spreadsheet or choose_spreadsheet()
    if spreadsheet is None:
        print("No spreadsheet selected. Nothing was copied.")
        return 1
    if not spreadsheet.is_file():
        print(f"ERROR: Spreadsheet was not found: {spreadsheet}")
        return 2
    return organise(spreadsheet, args.source_root, args.output_root)


if __name__ == "__main__":
    sys.exit(main())
